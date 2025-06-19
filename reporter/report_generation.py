import os
from datetime import datetime
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, Reference

# 1. 從新的位置匯入資料結構
from common.data_structures import AnalysisResult


class ReportGenerationAgent:
    def __init__(self, output_dir: str = "output"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def _apply_conditional_formatting(self, ws):
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # 根據 "過時狀態" 欄位 (B欄) 進行格式化
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            status_cell = row[1] # B欄
            if "❌" in status_cell.value:
                for cell in row:
                    cell.fill = red_fill
            elif "⚠️" in status_cell.value:
                for cell in row:
                    cell.fill = yellow_fill
            elif "✅" in status_cell.value:
                for cell in row:
                    cell.fill = green_fill

    def _create_dashboard(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("儀表板", 0) # Insert as the first sheet

        # --- Status Summary ---
        status_counts = df["過時狀態"].value_counts().reset_index()
        status_counts.columns = ["狀態", "網站數量"]
        
        for r_idx, row in enumerate(dataframe_to_rows(status_counts, index=False, header=True), 1):
            ws.append(row)

        # --- Pie Chart ---
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=len(status_counts) + 1)
        data = Reference(ws, min_col=2, min_row=1, max_row=len(status_counts) + 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "網站狀態分佈圖"
        ws.add_chart(pie, "D2")

        # --- Styling ---
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for cell in ws["1:1"]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15


    def generate(self, results: List[AnalysisResult]):
        data = [
            {
                "網站 URL": res.url,
                "過時狀態": res.status,
                "過時分數 (越高越舊)": res.score,
                "偵測到的最後更新日": res.last_updated,
                "AI 分析說明": res.notes,
                "失效連結列表": res.broken_links_summary,
            }
            for res in results
        ]
        df = pd.DataFrame(data)

        today = datetime.now().strftime("%Y%m%d")
        output_path = os.path.join(self.output_dir, f"report_{today}.xlsx")

        # Use openpyxl to create a more advanced report
        wb = Workbook()
        ws = wb.active
        ws.title = "詳細報告"

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # --- 格式化 --- 
        self._format_header(ws)
        self._apply_conditional_formatting(ws)
        self._auto_fit_columns(ws)

        # --- 儀表板 --- 
        self._create_dashboard(wb, df)

        wb.save(output_path)
        return output_path

    def _format_header(self, ws):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_fit_columns(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            # 為長文字欄位設定最大寬度
            if column in ['E', 'F']: # AI 分析說明 & 失效連結
                 ws.column_dimensions[column].width = min(adjusted_width, 80)
                 for cell in ws[column]:
                     cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                 ws.column_dimensions[column].width = adjusted_width
